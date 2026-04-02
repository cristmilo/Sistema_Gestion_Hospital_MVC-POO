# utils/exportar.py
# ─────────────────────────────────────────────
# Funciones para exportar datos a Excel y PDF.
# Se importan en las views para los botones
# de exportación de cada módulo.
# ─────────────────────────────────────────────

from tkinter import filedialog, messagebox
from openpyxl import Workbook
from fpdf import FPDF

def exportar_excel(titulo, encabezados, filas):
    """Genera un archivo .xlsx con los datos recibidos."""
    ruta = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel", "*.xlsx")],
        title="Guardar Excel"
    )
    if not ruta:
        return
    wb = Workbook()
    ws = wb.active
    ws.title = titulo
    ws.append(encabezados)
    for fila in filas:
        ws.append(list(fila))
    wb.save(ruta)
    messagebox.showinfo("Excel", f"Guardado en:\n{ruta}")

def exportar_pdf(titulo, encabezados, filas):
    """Genera un archivo .pdf con formato de tabla."""
    ruta = filedialog.asksaveasfilename(
        defaultextension=".pdf",
        filetypes=[("PDF", "*.pdf")],
        title="Guardar PDF"
    )
    if not ruta:
        return
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", "B", 14)
    pdf.cell(0, 10, titulo, ln=True, align="C")
    pdf.ln(4)
    ancho = 190 // len(encabezados)
    pdf.set_font("Arial", "B", 10)
    for enc in encabezados:
        pdf.cell(ancho, 8, str(enc), border=1, align="C")
    pdf.ln()
    pdf.set_font("Arial", "", 9)
    for fila in filas:
        for celda in fila:
            pdf.cell(ancho, 7, str(celda), border=1)
        pdf.ln()
    pdf.output(ruta)
    messagebox.showinfo("PDF", f"Guardado en:\n{ruta}")
